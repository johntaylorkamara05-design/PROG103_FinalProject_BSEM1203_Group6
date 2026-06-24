import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Initialize Workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Project Gantt Chart"

# Ensure grid lines are visible
ws.views.sheetView[0].showGridLines = True

# Define Colors and Fonts
font_family = "Tahoma"
title_font = Font(name=font_family, size=14, bold=True, color="FFFFFF")
header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
regular_font = Font(name=font_family, size=11)
bold_font = Font(name=font_family, size=11, bold=True)

dark_blue_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
light_blue_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
green_fill = PatternFill(start_color="2ECC71", end_color="2ECC71", fill_type="solid")

thin_border = Border(
    left=Side(style='thin', color='D3D3D3'),
    right=Side(style='thin', color='D3D3D3'),
    top=Side(style='thin', color='D3D3D3'),
    bottom=Side(style='thin', color='D3D3D3')
)

# Set Up Headings
ws.merge_cells("A1:M1")
ws['A1'] = "PROG 102 - PRINCIPLES OF SOFTWARE ENGINEERING GANTT CHART"
ws['A1'].font = title_font
ws['A1'].fill = dark_blue_fill
ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 40

# Column Headers
headers = ["Task Description", "W4", "W5", "W6", "W7", "W8", "W9", "W10", "W11", "Status", "Assigned To"]
for col_num, header_title in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col_num)
    cell.value = header_title
    cell.font = header_font
    cell.fill = dark_blue_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[3].height = 25

# Task Data Sets (Task, Start_Col, End_Col, Status)
tasks = [
    ("1. Requirements & SDG Alignment Exploration", 2, 3, "Completed"),
    ("2. UI/UX Prototype Design (10 Figma Frames)", 4, 5, "Completed"),
    ("3. System Architecture & UML Blueprinting", 6, 7, "Completed"),
    ("4. Technical Writing & License Attribution", 8, 8, "Completed"),
    ("5. GitHub Repository Push & PDF Submission", 9, 9, "Completed")
]

# Populate Rows
for idx, (task_text, start, end, status) in enumerate(tasks, 4):
    ws.row_dimensions[idx].height = 22

    # Task Name
    cell_task = ws.cell(row=idx, column=1, value=task_text)
    cell_task.font = regular_font
    cell_task.border = thin_border

    # Fill Timeline blocks
    for col in range(2, 10):
        cell_block = ws.cell(row=idx, column=col)
        cell_block.border = thin_border
        if start <= col <= end:
            cell_block.fill = light_blue_fill

    # Status
    cell_status = ws.cell(row=idx, column=10, value=status)
    cell_status.font = bold_font
    cell_status.fill = green_fill
    cell_status.alignment = Alignment(horizontal="center")
    cell_status.border = thin_border

    # Assigned To (Solo Lead)
    cell_assign = ws.cell(row=idx, column=11, value="John W. Taylor-Kamara")
    cell_assign.font = regular_font
    cell_assign.alignment = Alignment(horizontal="center")
    cell_assign.border = thin_border

# Auto-fit Column Widths
for col in ws.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    col_letter = openpyxl.utils.get_column_letter(col[0].column)
    ws.column_dimensions[col_letter].width = max(max_len + 3, 7)
ws.column_dimensions['A'].width = 45

# Save Output
wb.save("Project_Gantt_Chart.xlsx")
print("Success! 'Project_Gantt_Chart.xlsx' has been generated smoothly.")